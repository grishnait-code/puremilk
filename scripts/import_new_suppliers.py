"""
Импорт данных из новых файлов Quality Monitor в PostgreSQL.

Поддерживает два формата:
  - Старый (как Гатчинское/Новоладожское): 32 колонки, данные с row 35
  - Новый (2024): 33 колонки (+ИВ после AB), данные с row ~14-20

Автоматически определяет:
  - Начало строк с данными (первая ячейка col A = datetime)
  - Тип колонок (есть ли столбец ИВ после AB)

Использование:
    python scripts/import_new_suppliers.py

Или для конкретного файла:
    python scripts/import_new_suppliers.py --file "путь/к/файлу.xlsx" --enterprise "Название" --short "Краткое"

Опции:
    --clear   Удалить существующие поставки предприятия перед импортом
"""

import argparse
import os
import sys
import datetime

import openpyxl
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))
from app.models import Enterprise, Delivery, QualityResult, Base  # noqa

DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "postgresql://postgres:postgres@localhost:5433/quality_monitor",
)

# --- Определение файлов для импорта ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
DATA_DIR = os.path.join(os.path.dirname(PROJECT_DIR), "Данные для добавления в базу")

FILES = [
    {
        "file": os.path.join(DATA_DIR, "АО Оскольское молоко Quality Monitor.xlsx"),
        "enterprise": "АО Оскольское молоко",
        "short": "Оскольское молоко",
    },
    {
        "file": os.path.join(DATA_DIR, "МТК Вереск Quality Monitor.xlsx"),
        "enterprise": "МТК Вереск",
        "short": "Вереск",
    },
    {
        "file": os.path.join(DATA_DIR, "МТК Ромашка Quality Monitor.xlsx"),
        "enterprise": "МТК Ромашка",
        "short": "Ромашка",
    },
    {
        "file": os.path.join(DATA_DIR, "ЭНА ЖК Бобров-1 Quality Monitor.xlsx"),
        "enterprise": "ЭНА ЖК Бобров-1",
        "short": "Бобров-1",
    },
    {
        "file": os.path.join(DATA_DIR, "ЭНА ЖК Добрино Quality Monitor.xlsx"),
        "enterprise": "ЭНА ЖК Добрино",
        "short": "Добрино",
    },
]

# Маппинг колонок для старого формата (как Гатчинское, 0-based)
COL_OLD = {
    "date":               0,
    "weight_kg":          1,
    "weight_recalc_kg":   2,
    "grade_E_kg":         3,
    "grade_I_kg":         4,
    "grade_II_kg":        5,
    "grade_out_kg":       6,
    "grade_error_kg":     7,
    "grade_E_final_kg":   8,
    "has_antibiotics":    9,
    "temperature_lab":    10,
    "temperature_std":    11,
    "organoleptic_lab":   12,
    "organoleptic_std":   13,
    "scc":                14,
    "bact_count_lab":     15,
    "bact_count_std":     16,
    "freeze_point_lab":   17,
    "freeze_point_std":   18,
    "fat_pct":            19,
    "protein_pct":        20,
    "lactose_pct":        21,
    "snf_pct":            22,
    "density":            23,
    "alcohol_pct":        24,
    "acidity":            25,
    "ph_lab":             26,
    "ph_std":             27,
    "coliforms":          28,
    "fatty_acids":        29,
    "urea":               30,
    "clostridium_spores": 31,
}

# Маппинг для нового формата (+ИВ после AB, сдвиг с col 11)
COL_NEW = dict(COL_OLD)
COL_NEW.update({
    "temperature_lab":    11,
    "temperature_std":    12,
    "organoleptic_lab":   13,
    "organoleptic_std":   14,
    "scc":                15,
    "bact_count_lab":     16,
    "bact_count_std":     17,
    "freeze_point_lab":   18,
    "freeze_point_std":   19,
    "fat_pct":            20,
    "protein_pct":        21,
    "lactose_pct":        22,
    "snf_pct":            23,
    "density":            24,
    "alcohol_pct":        25,
    "acidity":            26,
    "ph_lab":             27,
    "ph_std":             28,
    "coliforms":          29,
    "fatty_acids":        30,
    "urea":               31,
    "clostridium_spores": 32,
})


def parse_num(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip().replace(",", ".").replace("—", "").replace("-", "")
        if not v:
            return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def detect_format(ws):
    """
    Определяет:
      - first_data_row: номер первой строки с данными (1-based)
      - col_map: маппинг колонок (COL_OLD или COL_NEW)
    """
    first_data_row = None
    for r in range(1, 200):
        val = ws.cell(r, 1).value
        if isinstance(val, datetime.datetime):
            first_data_row = r
            break
    if first_data_row is None:
        raise ValueError("Не найдена ни одна строка с датой в колонке A")

    # Ищем строку-заголовок уровня 2 (с 'ПО'/'Т001-2'/'AB'/'ИВ')
    # Обычно за 4 строки до первой даты
    has_iv = False
    for r in range(max(1, first_data_row - 6), first_data_row):
        row_vals = {c: ws.cell(r, c).value for c in range(1, 20)}
        # Ищем признак нового формата: col 11 = 'ИВ' или 'ИB'
        v11 = row_vals.get(11, None)
        if isinstance(v11, str) and v11.strip().upper() in ('ИВ', 'ИB', 'ИB', 'IB'):
            has_iv = True
            break
        # Альтернатива: col 11 = 'Наличие' (второй столбец Наличие)
        v11_15 = row_vals.get(15)
        v11_header = row_vals.get(11)
        if isinstance(v11_header, str) and 'наличие' in v11_header.lower():
            has_iv = True
            break

    col_map = COL_NEW if has_iv else COL_OLD
    fmt = "новый (+ИВ)" if has_iv else "старый"
    print(f"  Формат: {fmt}, данные с строки {first_data_row}")
    return first_data_row, col_map


def import_file(file_path, enterprise_name, short_name, clear, db):
    if not os.path.exists(file_path):
        print(f"  ОШИБКА: файл не найден: {file_path}")
        return

    wb = openpyxl.load_workbook(file_path, data_only=True)
    if "Quality" not in wb.sheetnames:
        print(f"  ОШИБКА: лист 'Quality' не найден в файле")
        return
    ws = wb["Quality"]

    first_data_row, COL = detect_format(ws)

    # Получить или создать предприятие
    enterprise = db.query(Enterprise).filter(
        Enterprise.name == enterprise_name
    ).first()
    if not enterprise:
        enterprise = Enterprise(name=enterprise_name, short_name=short_name)
        db.add(enterprise)
        db.flush()
        print(f"  Создано предприятие: {enterprise_name} (id={enterprise.id})")
    else:
        print(f"  Предприятие найдено: {enterprise_name} (id={enterprise.id})")

    if clear:
        deleted = db.query(Delivery).filter(
            Delivery.enterprise_id == enterprise.id
        ).delete()
        db.flush()
        print(f"  Удалено {deleted} старых поставок")

    imported = 0
    skipped = 0

    for r in range(first_data_row, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 40)]

        # Дата
        date_val = row[COL["date"]]
        if date_val is None:
            continue
        if isinstance(date_val, datetime.datetime):
            delivery_date = date_val.date()
        elif isinstance(date_val, str):
            try:
                delivery_date = datetime.datetime.strptime(
                    date_val.strip(), "%d.%m.%Y"
                ).date()
            except ValueError:
                continue
        else:
            continue

        # Вес
        weight_kg = parse_num(row[COL["weight_kg"]])
        if weight_kg is None:
            continue

        # Проверка на дубликат
        existing = db.query(Delivery).filter(
            Delivery.enterprise_id == enterprise.id,
            Delivery.delivery_date == delivery_date,
        ).first()
        if existing and not clear:
            skipped += 1
            continue

        # Антибиотики: 0 = нет, 1/True = да
        ab_val = row[COL["has_antibiotics"]]
        has_ab = bool(ab_val) if ab_val not in (None, 0, 0.0) else False

        # Точка замерзания: конвертируем ×(-1000) если значение выглядит как -0.5xx
        def parse_freeze(v):
            n = parse_num(v)
            if n is None:
                return None
            # Если значение отрицательное (напр. -0.544) — конвертируем в тысячные
            if -1.0 < n < 0:
                return round(n * -1000, 1)
            return n

        delivery = Delivery(
            enterprise_id=enterprise.id,
            delivery_date=delivery_date,
            weight_kg=weight_kg,
            weight_recalc_kg=parse_num(row[COL["weight_recalc_kg"]]),
            grade_E_kg=parse_num(row[COL["grade_E_kg"]]),
            grade_I_kg=parse_num(row[COL["grade_I_kg"]]),
            grade_II_kg=parse_num(row[COL["grade_II_kg"]]),
            grade_out_kg=parse_num(row[COL["grade_out_kg"]]),
            grade_error_kg=parse_num(row[COL["grade_error_kg"]]),
            grade_E_final_kg=parse_num(row[COL["grade_E_final_kg"]]),
            has_antibiotics=has_ab,
        )
        db.add(delivery)
        db.flush()

        qr = QualityResult(
            delivery_id=delivery.id,
            temperature_lab=parse_num(row[COL["temperature_lab"]]),
            temperature_std=parse_num(row[COL["temperature_std"]]),
            organoleptic_lab=parse_num(row[COL["organoleptic_lab"]]),
            organoleptic_std=parse_num(row[COL["organoleptic_std"]]),
            scc=parse_num(row[COL["scc"]]),
            bact_count_lab=parse_num(row[COL["bact_count_lab"]]),
            bact_count_std=parse_num(row[COL["bact_count_std"]]),
            freeze_point_lab=parse_freeze(row[COL["freeze_point_lab"]]),
            freeze_point_std=parse_freeze(row[COL["freeze_point_std"]]),
            fat_pct=parse_num(row[COL["fat_pct"]]),
            protein_pct=parse_num(row[COL["protein_pct"]]),
            lactose_pct=parse_num(row[COL["lactose_pct"]]),
            snf_pct=parse_num(row[COL["snf_pct"]]),
            density=parse_num(row[COL["density"]]),
            alcohol_pct=parse_num(row[COL["alcohol_pct"]]),
            acidity=parse_num(row[COL["acidity"]]),
            ph_lab=parse_num(row[COL["ph_lab"]]),
            ph_std=parse_num(row[COL["ph_std"]]),
            coliforms=parse_num(row[COL["coliforms"]]),
            fatty_acids=parse_num(row[COL["fatty_acids"]]),
            urea=parse_num(row[COL["urea"]]),
            clostridium_spores=parse_num(row[COL["clostridium_spores"]]),
        )
        db.add(qr)
        imported += 1

    db.commit()
    print(f"  Импортировано: {imported} партий, пропущено дубликатов: {skipped}")


def main():
    parser = argparse.ArgumentParser(
        description="Импорт новых поставщиков из Quality Monitor xlsx"
    )
    parser.add_argument("--file", default=None, help="Путь к конкретному xlsx-файлу")
    parser.add_argument("--enterprise", default=None, help="Полное название предприятия")
    parser.add_argument("--short", default=None, help="Краткое название")
    parser.add_argument("--clear", action="store_true",
                        help="Удалить старые поставки перед импортом")
    args = parser.parse_args()

    db_url = DATABASE_URL.replace(
        "postgresql://", "postgresql+psycopg://"
    ) if "psycopg2" not in DATABASE_URL and "+psycopg" not in DATABASE_URL else DATABASE_URL

    engine = create_engine(db_url)
    Base.metadata.create_all(bind=engine)
    Session = sessionmaker(bind=engine)
    db = Session()

    try:
        if args.file:
            # Конкретный файл
            files = [{
                "file": args.file,
                "enterprise": args.enterprise or os.path.basename(args.file),
                "short": args.short or args.enterprise or os.path.basename(args.file),
            }]
        else:
            # Все 5 новых поставщиков
            files = FILES

        for entry in files:
            print(f"\n{'='*60}")
            print(f"Файл: {os.path.basename(entry['file'])}")
            print(f"Предприятие: {entry['enterprise']}")
            import_file(
                entry["file"],
                entry["enterprise"],
                entry["short"],
                args.clear,
                db,
            )

        print(f"\n{'='*60}")
        print("Импорт завершён.")

    except Exception as e:
        db.rollback()
        print(f"\nОШИБКА: {e}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
