"""
Импорт данных из файлов Quality Monitor (*.xlsx) в PostgreSQL.
Каждая строка листа Quality (начиная с row 35) = одна партия молока.

Использование:
    python scripts/import_xlsx.py \
        --file "Novoladozhskiy 2. Quality Monitor.xlsx" \
        --enterprise "Новоладожское" \
        --short "Новоладожское"

Опции:
    --clear   Удалить существующие поставки предприятия перед импортом
"""

import argparse
import os
import sys
import datetime

import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))
from app.models import Enterprise, Delivery, QualityResult, Base  # noqa

DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "postgresql://postgres:postgres@localhost:5433/quality_monitor",
)

# Маппинг колонок листа Quality (индексы с 0)
# Row 30-34: заголовки и нормативы — пропускаем
# Row 35+: партии
COL = {
    "date":             0,
    "weight_kg":        1,
    "weight_recalc_kg": 2,
    "grade_E_kg":       3,
    "grade_I_kg":       4,
    "grade_II_kg":      5,
    "grade_out_kg":     6,
    "grade_error_kg":   7,
    "grade_E_final_kg": 8,
    "has_antibiotics":  9,
    "temperature_lab":  10,
    "temperature_std":  11,
    "organoleptic_lab": 12,
    "organoleptic_std": 13,
    "scc":              14,
    "bact_count_lab":   15,
    "bact_count_std":   16,
    "freeze_point_lab": 17,
    "freeze_point_std": 18,
    "fat_pct":          19,
    "protein_pct":      20,
    "lactose_pct":      21,
    "snf_pct":          22,
    "density":          23,
    "alcohol_pct":      24,
    "acidity":          25,
    "ph_lab":           26,
    "ph_std":           27,
    "coliforms":        28,
    "fatty_acids":      29,
    "urea":             30,
    "clostridium_spores": 31,
}


def parse_num(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip().replace(",", ".").replace("-", "")
        if not v:
            return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def import_file(xlsx_path: str, enterprise_name: str, short_name: str,
                clear: bool, db):
    # Находим или создаём предприятие
    enterprise = db.query(Enterprise).filter(
        Enterprise.name == enterprise_name
    ).first()
    if not enterprise:
        enterprise = Enterprise(name=enterprise_name, short_name=short_name)
        db.add(enterprise)
        db.flush()
        print(f"Создано предприятие: {enterprise_name} (id={enterprise.id})")
    else:
        print(f"Предприятие найдено: {enterprise_name} (id={enterprise.id})")

    # Очищаем старые данные если нужно
    if clear:
        existing = db.query(Delivery).filter(
            Delivery.enterprise_id == enterprise.id
        ).all()
        for d in existing:
            db.query(QualityResult).filter(
                QualityResult.delivery_id == d.id
            ).delete()
        db.query(Delivery).filter(
            Delivery.enterprise_id == enterprise.id
        ).delete()
        db.flush()
        print(f"Удалено {len(existing)} старых записей")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if "Quality" not in wb.sheetnames:
        print("Лист 'Quality' не найден!")
        return

    ws = wb["Quality"]
    imported = 0
    skipped = 0

    # Партии начинаются с строки 35 (индекс 34 при min_row=35)
    for row in ws.iter_rows(min_row=35, values_only=True):
        # Первая колонка должна быть датой
        delivery_date = parse_date(row[COL["date"]])
        if delivery_date is None:
            skipped += 1
            continue

        weight_kg = parse_num(row[COL["weight_kg"]])

        # Пропускаем строки без веса
        if weight_kg is None or weight_kg == 0:
            skipped += 1
            continue

        # Проверяем дубликат
        existing = db.query(Delivery).filter(
            Delivery.enterprise_id == enterprise.id,
            Delivery.delivery_date == delivery_date,
        ).first()
        if existing and not clear:
            skipped += 1
            continue

        # Антибиотики: 0 = нет, 1 = да
        ab_val = row[COL["has_antibiotics"]]
        has_ab = bool(ab_val) if ab_val not in (None, 0, 0.0) else False

        delivery = Delivery(
            enterprise_id=enterprise.id,
            delivery_date=delivery_date,
            weight_kg=weight_kg,
            weight_recalc_kg=parse_num(row[COL["weight_recalc_kg"]]),
            grade_E_kg=parse_num(row[COL["grade_E_kg"]]),
            grade_I_kg=parse_num(row[COL["grade_I_kg"]]),
            grade_II_kg=parse_num(row[COL["grade_II_kg"]]),
            grade_out_kg=parse_num(row[COL["grade_out_kg"]]),
            grade_error_kg=parse_num(row[COL["grade_error_kg"]]),
            grade_E_final_kg=parse_num(row[COL["grade_E_final_kg"]]),
            has_antibiotics=has_ab,
        )
        db.add(delivery)
        db.flush()

        qr = QualityResult(
            delivery_id=delivery.id,
            temperature_lab=parse_num(row[COL["temperature_lab"]]),
            temperature_std=parse_num(row[COL["temperature_std"]]),
            organoleptic_lab=parse_num(row[COL["organoleptic_lab"]]),
            organoleptic_std=parse_num(row[COL["organoleptic_std"]]),
            scc=parse_num(row[COL["scc"]]),
            bact_count_lab=parse_num(row[COL["bact_count_lab"]]),
            bact_count_std=parse_num(row[COL["bact_count_std"]]),
            freeze_point_lab=parse_num(row[COL["freeze_point_lab"]]),
            freeze_point_std=parse_num(row[COL["freeze_point_std"]]),
            fat_pct=parse_num(row[COL["fat_pct"]]),
            protein_pct=parse_num(row[COL["protein_pct"]]),
            lactose_pct=parse_num(row[COL["lactose_pct"]]),
            snf_pct=parse_num(row[COL["snf_pct"]]),
            density=parse_num(row[COL["density"]]),
            alcohol_pct=parse_num(row[COL["alcohol_pct"]]),
            acidity=parse_num(row[COL["acidity"]]),
            ph_lab=parse_num(row[COL["ph_lab"]]),
            ph_std=parse_num(row[COL["ph_std"]]),
            coliforms=parse_num(row[COL["coliforms"]]),
            fatty_acids=parse_num(row[COL["fatty_acids"]]),
            urea=parse_num(row[COL["urea"]]),
            clostridium_spores=parse_num(row[COL["clostridium_spores"]]),
        )
        db.add(qr)
        imported += 1

    db.commit()
    print(f"Импортировано: {imported} партий, пропущено: {skipped}")


def main():
    parser = argparse.ArgumentParser(
        description="Импорт партий молока из Quality Monitor xlsx"
    )
    parser.add_argument("--file", required=True, help="Путь к xlsx-файлу")
    parser.add_argument("--enterprise", required=True,
                        help="Полное название предприятия")
    parser.add_argument("--short", default=None,
                        help="Краткое название предприятия")
    parser.add_argument("--clear", action="store_true",
                        help="Удалить старые поставки перед импортом")
    args = parser.parse_args()

    db_url = DATABASE_URL.replace(
        "postgresql://", "postgresql+psycopg2://"
    ) if "psycopg2" not in DATABASE_URL else DATABASE_URL

    engine = create_engine(db_url)
    Base.metadata.create_all(bind=engine)
    Session = sessionmaker(bind=engine)
    db = Session()

    try:
        import_file(
            args.file,
            args.enterprise,
            args.short or args.enterprise,
            args.clear,
            db
        )
    except Exception as e:
        db.rollback()
        print(f"Ошибка: {e}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
